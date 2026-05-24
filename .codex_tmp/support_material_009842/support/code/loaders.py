"""读取附件 1-8 的 Excel 数据，把它们转为干净的 numpy 数组。

约定：所有 24 小时序列长度都是 24，索引 h 对应时段 [h, h+1)；
功率单位 MW，电量单位 MWh，电价单位 元/kWh。
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

SUPPORT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = SUPPORT_DIR / "data"
LEGACY_DATA_DIR = Path(__file__).resolve().parent / "A题"

FN_LOAD = "附件1：园区典型日常规电负荷标幺功率曲线.xlsx"
FN_TYPICAL = "附件2:典型日风电、光伏标幺功率表.xlsx".replace(":", "：")
FN_WIND_SCN = "附件3：园区6种场景的风电标幺功率表.xlsx"
FN_PV_SCN = "附件4：园区4种场景的光伏标幺功率表.xlsx"


def _read_24h_matrix(fname: str, ncols: int) -> list[list[float]]:
    """读 (25 行 × (1+ncols) 列) 的表，扔掉表头行 + 时段列，返回 (24, ncols) float。"""
    path = DATA_DIR / fname
    if not path.exists():
        path = LEGACY_DATA_DIR / fname
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    arr: list[list[float]] = []
    for row in ws.iter_rows(min_row=2, max_row=25, min_col=2, max_col=1 + ncols, values_only=True):
        arr.append([float(value) for value in row])
    assert len(arr) == 24 and all(len(row) == ncols for row in arr), (
        f"{fname}: expected (24, {ncols}), got ({len(arr)}, {len(arr[0]) if arr else 0})"
    )
    return arr


def load_typical_load() -> list[float]:
    """附件 1：常规负荷标幺，shape (24,)。"""
    return [row[0] for row in _read_24h_matrix(FN_LOAD, 1)]


def load_typical_wind_pv() -> tuple[list[float], list[float]]:
    """附件 2：典型日风电、光伏标幺，两个 (24,) 向量。"""
    arr = _read_24h_matrix(FN_TYPICAL, 2)
    return [row[0] for row in arr], [row[1] for row in arr]


def load_wind_scenarios() -> list[list[float]]:
    """附件 3：6 个风电场景标幺，shape (24, 6)。"""
    return _read_24h_matrix(FN_WIND_SCN, 6)


def load_pv_scenarios() -> list[list[float]]:
    """附件 4：4 个光伏场景标幺，shape (24, 4)。"""
    return _read_24h_matrix(FN_PV_SCN, 4)


if __name__ == "__main__":
    print("load:", len(load_typical_load()))
    w, p = load_typical_wind_pv()
    print("wind:", len(w), "pv:", len(p))
    print("wind scenarios:", len(load_wind_scenarios()), "x", len(load_wind_scenarios()[0]))
    print("pv scenarios:", len(load_pv_scenarios()), "x", len(load_pv_scenarios()[0]))
